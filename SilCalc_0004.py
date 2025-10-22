#!/usr/bin/env python
# coding: utf-8

# In[1]:


# History:
# 0001 - 10.07.25: Initial version with stable data acquisition from C/E matrix
# 0002 - 11.07.25: Addition of E/E overview class, including method get_pdm_codes()
# 0003 - 25.07.25: Addition of class for acquisition of FuSa data


# In[1]:


import openpyxl
import re
import pandas as pd
import os
import yaml
import warnings
# In[2]:


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

with open("config.yaml", "r") as file:
    config = yaml.safe_load(file)


# Class for storage of component related data
class Component:
    def __init__(self, pid_code='', pdm_code='', bmk_code='', pfh_avg=-1, pfd_avg=-1, sys_cap=-1):
        self.pid_code = pid_code
        self.pdm_code = pdm_code
        self.bmk_code = bmk_code
        self.pfh_avg = pfh_avg
        self.pfd_avg = pfd_avg
        self.sys_cap = sys_cap


# In[3]:


# Class for handling of FuSa related data

class FuSa:
    def __init__(self, config_path="config.yaml"):
        # Lade YAML-Konfiguration
        with open(config_path, "r") as file:
            config = yaml.safe_load(file)

        fusa_config = config["fusa"]

        # Setze Konfigurationswerte
        self.path = os.path.join(os.getcwd(), fusa_config["path"])
        self.col_name_comp_id = fusa_config["col_name_comp_id"]
        self.col_names_vs_comp_properties = fusa_config["col_names_vs_comp_properties"]
        self.fusa_data = {}

    def get_fusa_data(self, dataframe):
        try:
            fusa_data = pd.read_csv(self.path, delimiter=';')
        except Exception as e:
            print(f"FuSa.get_fusa_data(): Unable to open file \"{self.path}\". Error: {e}")
            return {}

        for sifu in dataframe.itertuples(index=True):
            for sensor in sifu.sensors:
                for col_name_tmp, fusa_prop_tmp in self.col_names_vs_comp_properties.items():
                    fusa_val_tmp = fusa_data.loc[fusa_data[self.col_name_comp_id] == sensor.pdm_code, col_name_tmp]

                    if not fusa_val_tmp.empty:
                        setattr(sensor, fusa_prop_tmp, fusa_val_tmp.values[0])
                    else:
                        print(f"FuSa.get_fusa_data(): Failed to get {fusa_prop_tmp} for sensor {sensor.pid_code}")

            for actuator in sifu.actuators:
                comp_id = actuator.pdm_code if actuator.pdm_code else actuator.bmk_code

                for col_name_tmp, fusa_prop_tmp in self.col_names_vs_comp_properties.items():
                    fusa_val_tmp = fusa_data.loc[fusa_data[self.col_name_comp_id] == comp_id, col_name_tmp]

                    if not fusa_val_tmp.empty:
                        setattr(actuator, fusa_prop_tmp, fusa_val_tmp.values[0])
                    else:
                        print(f"FuSa.get_fusa_data(): Failed to get {fusa_prop_tmp} for actuator {actuator.pid_code}")

        self.fusa_data = fusa_data
        return fusa_data



# In[4]:



class EeOverview:
    def __init__(self, config_path="config.yaml"):
        # Lade YAML-Konfiguration
        with open(config_path, "r") as file:
            config = yaml.safe_load(file)

        ee_config = config["ee_overview"]
        patterns = config["patterns"]

        # Setze Konfigurationswerte
        self.path = os.path.join(os.getcwd(), ee_config["path"])
        self.modules_vs_sheets = ee_config["modules_vs_sheets"]
        self.pattern_PID_fcmos = ee_config["pattern_PID_fcmos"]
        self.sheets = ee_config["sheets"]
        self.col_number_pid_code = ee_config["col_number_pid_code"]
        self.col_number_pdm_code = ee_config["col_number_pdm_code"]
        self.row_number_start = ee_config["row_number_start"]

        # Setze reguläre Ausdrücke
        self.pattern_PID_designation = patterns["pid_designation"]
        self.pattern_PDM_designation = patterns["pdm_designation"]

    def get_pdm_codes(self, dataframe):
        pid_vs_pdm = {}

        try:
            workbook = openpyxl.load_workbook(self.path)
        except Exception as e:
            print(f"EeOverview.get_pdm_codes(): Unable to open file \"{self.path}\". Error: {e}")
            return pid_vs_pdm

        try:
            for sheet_name in self.sheets:
                sheet = workbook[sheet_name]

                for row in range(self.row_number_start, sheet.max_row):
                    pid_code = sheet.cell(row, self.col_number_pid_code).value
                    pdm_code = sheet.cell(row, self.col_number_pdm_code).value

                    if (pid_code and pdm_code and
                        re.match(self.pattern_PID_designation, pid_code) and
                        re.match(self.pattern_PDM_designation, pdm_code) and
                        pid_code not in pid_vs_pdm):
                        pid_vs_pdm[pid_code] = pdm_code

        except Exception as e:
            print(f"EeOverview.get_pdm_codes(): Error reading sheet \"{sheet_name}\" in file \"{self.path}\". Error: {e}")

        for sifu in dataframe.itertuples(index=True):
            for sensor in sifu.sensors:
                pdm_code_tmp = sensor.pdm_code
                pid_code_tmp = sensor.pid_code

                for key in self.pattern_PID_fcmos:
                    pid_code_tmp = re.sub(key, self.pattern_PID_fcmos[key], pid_code_tmp)

                if pid_code_tmp in pid_vs_pdm:
                    sensor.pdm_code = pid_vs_pdm[pid_code_tmp]
                elif pid_code_tmp:
                    print("EeOverview.get_pdm_codes(): PID code not found: " + pid_code_tmp)
                elif sensor.bmk_code:
                    pass
                else:
                    print("EeOverview.get_pdm_codes(): Unknown sensor in " + sifu.sifu_name)

            for actuator in sifu.actuators:
                pid_code_tmp = actuator.pid_code
                for key in self.pattern_PID_fcmos:
                    pid_code_tmp = re.sub(key, self.pattern_PID_fcmos[key], pid_code_tmp)

                if pid_code_tmp in pid_vs_pdm:
                    actuator.pdm_code = pid_vs_pdm[pid_code_tmp]
                elif pid_code_tmp:
                    print("EeOverview.get_pdm_codes(): PID code not found: " + pid_code_tmp)
                elif actuator.bmk_code:
                    pass
                else:
                    print("EeOverview.get_pdm_codes(): Unknown actuator in " + sifu.sifu_name)




# In[5]:



class CeMatrix:
    def __init__(self, config_path="config.yaml"):
        # Lade YAML-Konfiguration
        with open(config_path, "r") as file:
            config = yaml.safe_load(file)

        ce_config = config["ce_matrix"]
        patterns = config["patterns"]

        # Setze Konfigurationswerte
        self.path = os.path.join(os.getcwd(), ce_config["path"])
        self.sheet = ce_config["sheet"]
        self.col_number_state = ce_config["col_number_state"]
        self.col_number_sifu_name = ce_config["col_number_sifu_name"]
        self.col_number_sil_value = ce_config["col_number_sil_value"]
        self.col_number_criteria_safety_action = ce_config["col_number_criteria_safety_action"]
        self.col_number_demand_mode = ce_config["col_number_demand_mode"]
        self.col_number_safety_action = ce_config["col_number_safety_action"]
        self.row_number_start = ce_config["row_number_start"]
        self.sifu_state_active = ce_config["sifu_state_active"]
        self.terms_vs_actuators = ce_config["terms_vs_actuators"]

        # Setze reguläre Ausdrücke
        self.pattern_PID_designation = patterns["pid_designation"]
        self.pattern_contactor_designation = patterns["contactor_designation"]

    def get_content(self):
        try:
            workbook = openpyxl.load_workbook(self.path)
        except Exception as e:
            print(f"CeMatrix.get_content(): Unable to open file \"{self.path}\". Error: {e}")
            return pd.DataFrame()

        try:
            sheet = workbook[self.sheet]
        except Exception as e:
            print(f"CeMatrix.get_content(): Unable to open sheet \"{self.sheet}\" in file \"{self.path}\". Error: {e}")
            return pd.DataFrame()

        df = pd.DataFrame(columns=['sifu_name', 'sil_required', 'demand_mode_required', 'sensors', 'actuators', 'sil_value_calculated'])

        for row in range(self.row_number_start, sheet.max_row):
            sifu_status = sheet.cell(row, self.col_number_state).value

            if sifu_status == self.sifu_state_active:
                sifu_name = sheet.cell(row, self.col_number_sifu_name).value
                sifu_sil_value = sheet.cell(row, self.col_number_sil_value).value
                sifu_demand_mode = sheet.cell(row, self.col_number_demand_mode).value

                # Sensors
                criteria_cell = sheet.cell(row, self.col_number_criteria_safety_action).value or ""
                sifu_sensors_pid_codes = sorted(set(re.findall(self.pattern_PID_designation, criteria_cell)))
                sifu_sensors = [Component(pid_code=code) for code in sifu_sensors_pid_codes]

                # Actuators
                sifu_actuators_string = sheet.cell(row, self.col_number_safety_action).value or ""
                for term, actuator in self.terms_vs_actuators.items():
                    sifu_actuators_string = sifu_actuators_string.replace(term, actuator)

                sifu_actuators_pid_codes = sorted(set(re.findall(self.pattern_PID_designation, sifu_actuators_string)))
                sifu_actuators = [Component(pid_code=code) for code in sifu_actuators_pid_codes]

                sifu_actuators_bmk_codes = sorted(set(re.findall(self.pattern_contactor_designation, sifu_actuators_string)))
                sifu_actuators += [Component(bmk_code=code) for code in sifu_actuators_bmk_codes]

                df.loc[len(df)] = [sifu_name, sifu_sil_value, sifu_demand_mode, sifu_sensors, sifu_actuators, -1]

        if df.empty:
            print(f"CeMatrix.get_content(): No valid data found in sheet \"{self.sheet}\" of file \"{self.path}\".")
        else:
            print(f"CeMatrix.get_content(): Successfully acquired {df.shape[0]} entries.")

        return df



# In[6]:


ce_matrix = CeMatrix("config.yaml")
ee_overview = EeOverview("config.yaml")
fusa = FuSa("config.yaml")





# In[7]:


data = ce_matrix.get_content()
dict_pid_pdm = ee_overview.get_pdm_codes(data)
dings = fusa.get_fusa_data(data)


# In[14]:


for sifu in data.itertuples(index=True):
    print('\n'+sifu.sifu_name+': '+"SIL "+str(sifu.sil_required)+", "+sifu.demand_mode_required)
    for sensor in sifu.sensors:
        print('\t'+sensor.pid_code)
        print('\t\t'+"PFD Avg:\t"+str(sensor.pfd_avg))
        print('\t\t'+"PFH Avg:\t"+str(sensor.pfh_avg))
        print('\t\t'+"Sys. cap.:\t"+str(sensor.sys_cap))
    for actuator in sifu.actuators:
        if actuator.pid_code == '':
            print('\t'+actuator.bmk_code)    
        else:
            print('\t'+actuator.pid_code)
        print('\t\t'+"PFD Avg:\t"+str(actuator.pfd_avg))
        print('\t\t'+"PFH Avg:\t"+str(actuator.pfh_avg))
        print('\t\t'+"Sys. cap.:\t"+str(actuator.sys_cap))


